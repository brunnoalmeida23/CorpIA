# sharepoint.py
import requests
import io
import os
from config import DRIVE_ID

try:
    from docx import Document
    DOCX_AVAILABLE = True
except ImportError:
    DOCX_AVAILABLE = False

try:
    from openpyxl import load_workbook
    XLSX_AVAILABLE = True
except ImportError:
    XLSX_AVAILABLE = False

try:
    from PyPDF2 import PdfReader
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False

EXTENSOES_IGNORAR = ['.png', '.jpg', '.jpeg', '.gif', '.mp4', '.avi', '.mov', '.mp3', '.wav', '.exe', '.msi', '.zip', '.rar', '.7z', '.iso', '.img']

def listar_raiz_sharepoint(access_token):
    headers = {"Authorization": f"Bearer {access_token}"}
    url = f"https://graph.microsoft.com/v1.0/drives/{DRIVE_ID}/root/children"
    try:
        response = requests.get(url, headers=headers, timeout=30)
        if response.status_code == 401:
            return {"error": "Token expirado ou invalido"}
        response.raise_for_status()
        return response.json()
    except Exception as e:
        return {"error": str(e)}

def listar_pasta(access_token, item_id):
    headers = {"Authorization": f"Bearer {access_token}"}
    url = f"https://graph.microsoft.com/v1.0/drives/{DRIVE_ID}/items/{item_id}/children"
    try:
        response = requests.get(url, headers=headers, timeout=30)
        response.raise_for_status()
        return response.json()
    except Exception as e:
        return {"error": str(e)}

def buscar_todas_pastas(access_token):
    """Busca todas as pastas visiveis ao usuario no SharePoint (ate 2 niveis)"""
    pastas_encontradas = []
    raiz = listar_raiz_sharepoint(access_token)
    if "error" in raiz:
        return pastas_encontradas, raiz.get("error", "Erro desconhecido")
    
    itens = raiz.get("value", [])
    for item in itens:
        if "folder" in item:
            nome = item["name"]
            item_id = item["id"]
            pastas_encontradas.append({"nome": nome, "id": item_id, "caminho": nome})
            
            # Busca subpastas (nivel 1)
            try:
                sub_itens = listar_pasta(access_token, item_id)
                if "value" in sub_itens:
                    for sub_item in sub_itens["value"]:
                        if "folder" in sub_item:
                            sub_nome = sub_item["name"]
                            sub_id = sub_item["id"]
                            caminho = f"{nome}/{sub_nome}"
                            pastas_encontradas.append({
                                "nome": sub_nome,
                                "id": sub_id,
                                "caminho": caminho
                            })
                            
                            # Busca sub-subpastas (nivel 2)
                            try:
                                sub_sub_itens = listar_pasta(access_token, sub_id)
                                if "value" in sub_sub_itens:
                                    for sub_sub_item in sub_sub_itens["value"]:
                                        if "folder" in sub_sub_item:
                                            sub_sub_nome = sub_sub_item["name"]
                                            sub_sub_id = sub_sub_item["id"]
                                            pastas_encontradas.append({
                                                "nome": sub_sub_nome,
                                                "id": sub_sub_id,
                                                "caminho": f"{caminho}/{sub_sub_nome}"
                                            })
                            except:
                                pass
            except:
                pass
    
    return pastas_encontradas, None

def buscar_pasta_por_nome(access_token, nome_procurado):
    """Busca uma pasta especifica pelo nome (ignora pontos, espacos, maiusculas)"""
    print(f"[SHAREPOINT] Procurando pasta: '{nome_procurado}'...")
    
    pastas, erro = buscar_todas_pastas(access_token)
    if erro:
        return None, f"Erro: {erro}"
    
    # Limpa o nome procurado: minusculo, sem pontos, sem espacos
    nome_limpo = nome_procurado.lower().replace('.', '').replace(' ', '').replace('-', '').replace('_', '')
    
    # Primeiro tenta match exato
    for pasta in pastas:
        pasta_limpa = pasta['nome'].lower().replace('.', '').replace(' ', '').replace('-', '').replace('_', '')
        if nome_limpo == pasta_limpa:
            print(f"[SHAREPOINT] Pasta encontrada (exata): {pasta['caminho']}")
            return pasta, None
    
    # Depois tenta match parcial (contem)
    for pasta in pastas:
        pasta_limpa = pasta['nome'].lower().replace('.', '').replace(' ', '').replace('-', '').replace('_', '')
        if nome_limpo in pasta_limpa or pasta_limpa in nome_limpo:
            print(f"[SHAREPOINT] Pasta encontrada (parcial): {pasta['caminho']}")
            return pasta, None
    
    return None, f"Pasta '{nome_procurado}' nao encontrada."

def listar_arquivos_pasta(access_token, pasta_id):
    headers = {"Authorization": f"Bearer {access_token}"}
    url = f"https://graph.microsoft.com/v1.0/drives/{DRIVE_ID}/items/{pasta_id}/children"
    try:
        response = requests.get(url, headers=headers, timeout=30)
        if response.status_code == 401:
            return {"error": "Token expirado"}
        response.raise_for_status()
        return response.json()
    except Exception as e:
        return {"error": str(e)}

def baixar_arquivo(access_token, item_id):
    headers = {"Authorization": f"Bearer {access_token}"}
    try:
        url = f"https://graph.microsoft.com/v1.0/drives/{DRIVE_ID}/items/{item_id}"
        r = requests.get(url, headers=headers, timeout=30)
        r.raise_for_status()
        data = r.json()
        download_url = data.get("@microsoft.graph.downloadUrl")
        if not download_url:
            return None
        response = requests.get(download_url, timeout=60)
        response.raise_for_status()
        return response.content
    except:
        return None

def ler_arquivo_texto(conteudo_bytes):
    if not conteudo_bytes:
        return ""
    try:
        for encoding in ['utf-8', 'latin-1', 'cp1252']:
            try:
                return conteudo_bytes.decode(encoding)
            except:
                continue
        return conteudo_bytes.decode('utf-8', errors='ignore')
    except:
        return ""

def ler_arquivo_docx(conteudo_bytes):
    if not DOCX_AVAILABLE or not conteudo_bytes:
        return ""
    try:
        doc = Document(io.BytesIO(conteudo_bytes))
        return "\n".join([p.text for p in doc.paragraphs if p.text.strip()])
    except:
        return ""

def ler_arquivo_xlsx(conteudo_bytes):
    if not XLSX_AVAILABLE or not conteudo_bytes:
        return ""
    try:
        wb = load_workbook(io.BytesIO(conteudo_bytes), data_only=True)
        texto = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows(values=True):
                linha = [str(cell) for cell in row if cell is not None]
                if linha:
                    texto.append(" | ".join(linha))
        return "\n".join(texto)
    except:
        return ""

def ler_arquivo_pdf(conteudo_bytes):
    if not PDF_AVAILABLE or not conteudo_bytes:
        return ""
    try:
        reader = PdfReader(io.BytesIO(conteudo_bytes))
        return "\n".join([page.extract_text() for page in reader.pages if page.extract_text().strip()])
    except:
        return ""

def ler_arquivos_da_pasta_ti(access_token):
    """Retorna lista de pastas principais do usuario"""
    print("[SHAREPOINT] Buscando pastas do usuario...")
    
    pastas, erro = buscar_todas_pastas(access_token)
    if erro:
        return f"Erro ao acessar SharePoint: {erro}"
    if not pastas:
        return "Nenhuma pasta encontrada no seu SharePoint."
    
    # Pega so pastas principais (raiz)
    pastas_principais = []
    for p in pastas:
        if '/' not in p['caminho']:
            pastas_principais.append(p['caminho'])
    
    contexto = "PASTAS PRINCIPAIS DO SEU SHAREPOINT:\n"
    for nome in pastas_principais[:40]:
        contexto += f"- {nome}\n"
    
    if len(pastas_principais) > 40:
        contexto += f"\n... e mais **{len(pastas_principais) - 40} pastas**.\n"
    
    contexto += f"\nTotal: **{len(pastas_principais)} pastas** na raiz."
    contexto += "\n\nPergunte sobre uma pasta especifica para ver os arquivos (ex: 'O que tem na pasta TI?')."
    
    print(f"[SHAREPOINT] Contexto: {len(contexto)} caracteres")
    return contexto